from __future__ import annotations

import csv
import datetime
import time
from functools import wraps, singledispatch
from sys import getsizeof
from types import UnionType
from typing import Callable, ParamSpec, Concatenate, TypeVar, Type, Annotated, Generic, Optional, Tuple, Any, Union, \
    TypedDict
import openpyxl as op
import openpyxl.styles
from memory_profiler import profile

header = ['名次', '上次', 'aid', '标题', 'mid', 'up主', '投稿时间', '时长', '分P数', '播放增量', '弹幕增量', '评论增量',
          '收藏增量', '硬币增量', '分享增量',
          '点赞增量', 'Pt', '修正A', '修正B', '修正C', '长期入榜及期数', '收录', "引擎", '原创', "主榜", 'Last Pt',
          'rate', 'staff', '新曲排名', '新曲', '未授权搬运', '已删稿', 'HOT']
xlsx_header = ['名次', '上次', 'aid', '标题', 'mid', 'up主', '投稿时间', '时长', '分P数', '播放增量', '弹幕增量',
               '评论增量', '收藏增量', '硬币增量',
               '分享增量',
               '点赞增量', 'Pt', '修正A', '修正B', '修正C', 'Last Pt', 'rate', '长期入榜及期数', '新曲排名', 'staff',
               '原创', '引擎']
history_header = ['名次', '上次', 'aid', '标题', 'up主', '投稿时间', '时长', '分P数', '播放增量', '弹幕增量',
                  '评论增量', '收藏增量', '硬币增量',
                  '分享增量',
                  '点赞增量', 'Pt', '修正A', '修正B', '修正C', 'Last Pt', 'rate', 'staff']
xlsx_header_index = {i + 1: header[i] for i in range(len(header))}

int_data = ['aid', 'mid', '时长', '分P数', '播放增量', '弹幕增量', '评论增量', '收藏增量', '硬币增量', '分享增量',
            '点赞增量', '长期入榜及期数', '收录', '新曲排名', '已删稿']
float_data = ['Pt', '修正A', '修正B', '修正C']


#class Data_dict_required(TypedDict):
#    aid: int
#
#
# noinspection NonAsciiCharacters
#class Data_dict(Data_dict_required, total=False):
#    mid: int
#    时长: int
#    分P数: int
#    播放增量: int
#    弹幕增量: int
#    评论增量: int
#    收藏增量: int
#    硬币增量: int
#    分享增量: int
#    点赞增量: int
#    长期入榜及期数: int
#    收录: int
#    新曲排名: int
#    已删稿: int


def timer(func):
    def wrapper(*args, **kwargs):
        start_time = time.time()
        func(*args, **kwargs)
        print(f'cost:{time.time() - start_time}')

    return wrapper()


P = ParamSpec('P')
T = TypeVar("T")


def permission_access_decorator(func: Callable[Concatenate[str, P], T]) -> Callable[Concatenate[str, P], T]:
    @wraps(func)
    def wrapper(file: str, *args: P.args, **kwargs: P.kwargs) -> T:
        try:
            return func(file, *args, **kwargs)
        except PermissionError:
            print(f"{file}被占用，请关闭文件后按任意键重试")
            input()
            return wrapper(file, *args, **kwargs)

    return wrapper


@singledispatch
def time_covert(time_data: Any) -> Tuple[datetime.datetime, time.struct_time, str]:
    raise TypeError(f"不支持的时间类型{type(time_data)}")


@time_covert.register
def _(time_data: str) -> Tuple[datetime.datetime, time.struct_time, str]:
    struct_time: time.struct_time
    try:
        time.strptime(time_data, '%Y/%m/%d %H:%M:%S')
    except ValueError:
        try:
            struct_time = time.strptime(time_data, '%Y/%m/%d %H:%M')
        except ValueError:
            try:
                struct_time = time.strptime(time_data, '%Y/%m/%d %H:%M')
            except ValueError:
                struct_time = time.strptime(time_data, '%Y/%m/%d %H:%M:%S')
    else:
        raise TypeError(f"不支持的时间格式{time_data}")
    time_datetime = datetime.datetime.fromtimestamp(time.mktime(struct_time))
    return time_datetime, struct_time, time_data


@time_covert.register
def _(time_data: datetime.datetime) -> Tuple[datetime.datetime, time.struct_time, str]:
    time_str = time_data.strftime('%Y/%m/%d %H:%M')
    struct_time = time.strptime(time_str, '%Y/%m/%d %H:%M')
    return time_data, struct_time, time_str


Data_value_type = Union[int, float, str]


class Data:
    pub_time_datetime: datetime.datetime

    main_flag = 0  # 主榜/副榜 flag
    header_word_to_digit: dict[str, int] = {'名次': 0, '上次': 1, 'aid': 2, '标题': 3, 'mid': 4, 'up主': 5,
                                            '投稿时间': 6, '时长': 7, '分P数': 8, '播放增量': 9, '弹幕增量': 10,
                                            '评论增量': 11, '收藏增量': 12, '硬币增量': 13, '分享增量': 14,
                                            '点赞增量': 15,
                                            'Pt': 16, '修正A': 17, '修正B': 18, '修正C': 19, 'Last Pt': 20, 'rate': 21}
    header_digit_to_word: dict[int, str] = {value: key for key, value in header_word_to_digit.items()}
    int_data = ['aid', 'mid', '时长', '分P数', '播放增量', '弹幕增量', '评论增量', '收藏增量', '硬币增量', '分享增量',
                '点赞增量', '长期入榜及期数', '收录', '新曲排名', '已删稿']
    float_data = ['Pt', '修正A', '修正B', '修正C']

    @staticmethod
    def write_to_xlsx_wrapper(file_name: str, header: Optional[list['str'] | str] = None):
        if header is None:
            header = xlsx_header
        if header == 'history':
            header = history_header
        wb = op.Workbook()
        ws = wb.active
        for idx, key in enumerate(header):
            ws.cell(row=1, column=idx + 1).value = key
        line = 2

        def write_to_xlsx(self: Data):
            nonlocal line
            nonlocal ws
            nonlocal header
            font, pattern, border = self.xlsx_cell_style()
            if header is None:
                header = xlsx_header
            for idx, key in enumerate(header):
                try:
                    if key in ['修正A', '修正B', '修正C']:
                        ws.cell(row=line, column=idx + 1).value = round(float(self[key]), 3)
                        ws.cell(row=line, column=idx + 1).number_format = '0.000'
                    elif key == 'rate':
                        ws.cell(row=line, column=idx + 1).value = round(float(self[key]), 6)
                        ws.cell(row=line, column=idx + 1).number_format = '0.000%'
                    elif key in ['播放增量', '弹幕增量', '评论增量', '收藏增量', '硬币增量', '分享增量', '点赞增量',
                                 'Pt', 'Last Pt']:
                        ws.cell(row=line, column=idx + 1).value = int(float(self[key]))
                        ws.cell(row=line, column=idx + 1).number_format = '#,##0'
                    elif key == '投稿时间':
                        ws.cell(row=line, column=idx + 1).value = self[key]
                        ws.cell(row=line, column=idx + 1).number_format = 'yyyy/mm/dd hh:mm'
                    elif key == '新曲排名':
                        if int(float(self[key])) == 0:
                            ws.cell(row=line, column=idx + 1).value = None
                        else:
                            ws.cell(row=line, column=idx + 1).value = int(float(self[key]))
                            ws.cell(row=line, column=idx + 1).number_format = '000'
                    elif str(self[key]).isdigit():
                        ws.cell(row=line, column=idx + 1).value = int(self[key])
                    else:
                        ws.cell(row=line, column=idx + 1).value = self[key]
                except ValueError:
                    ws.cell(row=line, column=idx + 1).value = self[key]
                ws.cell(row=line, column=idx + 1).font = font
                ws.cell(row=line, column=idx + 1).fill = pattern
                ws.cell(row=line, column=idx + 1).border = border
            line += 1
            return

        #@profile
        def save_close():
            @permission_access_decorator
            def _save_close(file: str):
                nonlocal wb
                wb.save(file)
                wb.close()
                del wb
            _save_close(file_name)

        return write_to_xlsx, save_close

    @staticmethod
    def write_to_csv_wrapper(file_name: str,
                             *,
                             _header: list = header,
                             with_seconds: bool = False,
                             with_format: bool = False) \
            -> Tuple[Callable[[Data], None], Callable[[], None]]:
        #  一次性写入，覆盖原有数据

        @permission_access_decorator
        def __open(file: str):
            f = open(file_name, 'w+', encoding='utf-8-sig', newline='')
            writer = csv.DictWriter(f, fieldnames=_header)
            writer.writeheader()
            return f, writer

        f, writer = __open(file_name)

        def write_to_csv(self: Data) -> None:
            dict_to_write = {key: value for key, value in self.dict_.items() if key in _header}
            if with_seconds:
                dict_to_write['投稿时间'] = self.pub_time_datetime.strftime('%Y/%m/%d %H:%M:%S')
            else:
                dict_to_write['投稿时间'] = self.pub_time_datetime.strftime('%Y/%m/%d %H:%M')
            if with_format:
                for key in _header:
                    try:
                        if key in ['修正A', '修正B', '修正C']:
                            dict_to_write[key] = f'{float(self[key]):.3f}'
                        elif key == 'rate':
                            dict_to_write[key] = f'{float(self[key]) * 100:.3f}%'
                        elif key in ['播放增量', '弹幕增量', '评论增量', '收藏增量', '硬币增量', '分享增量', '点赞增量',
                                     'Pt', 'Last Pt']:
                            dict_to_write[key] = f'{int(float(self[key])):,}'
                        elif key == '新曲排名':
                            if int(float(self[key])) == 0:
                                dict_to_write[key] = ''
                            else:
                                dict_to_write[key] = f'{int(float(self[key])):03}'
                        elif str(self[key]).isdigit():
                            dict_to_write[key] = f'{int(self[key])}'
                        else:
                            dict_to_write[key] = f'{self[key]}'
                    except ValueError:
                        dict_to_write[key] = f'{self[key]}'
            writer.writerow(dict_to_write)

        #@profile
        def save_close() -> None:
            f.close()

        return write_to_csv, save_close

    def __init__(self, data: dict | list | tuple, file_type: str, file_header: list[str] = None,
                 required_keys: list[str] = None):
        # data_type必须为"xlsx"或“csv", data 是tuple, list, dict其中之一，前两者要求提供列索引(list)，
        self.dict_: dict[str, Data_value_type] = {}
        self.pub_time_datetime: datetime.datetime
        self.pub_time_time_struct: time.struct_time
        if required_keys is None:
            required_keys = ['aid']
        self.required_keys = required_keys
        if file_header is None:
            self.file_header = ['']
        else:
            self.file_header = file_header  # xlsx第i列列索引，从1开始
        if file_type == 'csv':
            if isinstance(data, dict):
                self.dict_ = data
            elif isinstance(data, (list, tuple)):
                if self.file_header == [''] or len(data) < len(self.file_header):
                    print("列索引错误")
                    input()
                    raise ValueError
                else:
                    self.dict_ = {self.file_header[i]: data[i] for i in range(len(self.file_header))}
            else:
                print("类型错误")
                input()
                raise TypeError
        elif file_type == 'xlsx':
            if self.file_header == [''] or len(data) > len(self.file_header) - 1:
                print("列索引不足")
                input()
                raise RuntimeError
            dict_: dict[str, Any] = {}
            for idx, k in enumerate(data):
                if key := self.file_header[idx + 1]:
                    dict_[key] = k.value
                else:
                    continue
                try:
                    if not k.fill is None:
                        if k.fill.start_color.rgb[2:] == 'FFFF00':
                            dict_['收录'] = 0
                        if k.fill.start_color.rgb[2:] in ['C00000', 'FF0000', '92D050']:
                            dict_['原创'] = '原创'
                            dict_['收录'] = 1
                        if k.fill.start_color.rgb[2:] == 'FFC000':
                            dict_['未授权搬运'] = '未授权搬运'
                            dict_['收录'] = 1
                except AttributeError or TypeError:
                    pass
                try:
                    if not k.border.bottom is None:
                        if k.border.bottom.style is not None:
                            # print(Data.main_flag)
                            if not Data.main_flag:
                                # print(k.value)
                                dict_['主榜'] = '主榜截止'  # 主榜到此截止
                                # print(self['title'])
                            else:
                                dict_['主榜'] = '副榜截止'  # 副榜截止
                except AttributeError:
                    pass
            self.dict_ = dict_
        else:
            print('文件格式错误')
            input()
            raise ValueError
        if not (isinstance(self['aid'], int) or (isinstance(self['aid'], str) and self['aid'].isdigit())):
            raise ValueError(f'缺少必要的aid')
        else:
            self.valid = True
        if '投稿时间' in self.dict_.keys():
            try:
                self.pub_time_datetime, self.pub_time_time_struct, self.dict_['投稿时间'] = time_covert(
                    self['投稿时间'])
            except TypeError as e:
                raise ValueError(f'{e}，在{self["aid"]}中')
        if '收录' in self.dict_.keys() and isinstance(self.dict_['收录'], str) and self.dict_['收录'].isdigit():
            self.dict_['收录'] = int(self.dict_['收录'])
        if '排名' in self.dict_.keys():
            self.dict_['名次'] = self.dict_['排名']
        for key in header + self.file_header:
            if key not in self.dict_.keys() or self.dict_[key] is None:
                self.dict_[key] = ''
        if '' in self.dict_.keys():
            del self.dict_['']
        _staff: Data_value_type = ''
        del_list = []
        for key in self.dict_.keys():
            if isinstance(key, str) and key.lower() == 'staff':
                _staff = self.dict_[key]
                del_list.append(key)
        for key in self.required_keys:
            if self.dict_.get(key) is None:
                raise ValueError(f'缺少必要的键: {key}')
        for key in del_list:
            del self.dict_[key]
        self.dict_['staff'] = _staff
        for key in Data.int_data:
            if key in self.dict_.keys() and isinstance(self.dict_[key], (str, float)) and self.dict_[key] not in ['',
                                                                                                                  '-',
                                                                                                                  '--']:
                self.dict_[key] = int(round(float(self.dict_[key]), 0))
        for key in Data.float_data:
            if key in self.dict_.keys() and isinstance(self.dict_[key], (str, float)) and self.dict_[key] not in ['',
                                                                                                                  '-',
                                                                                                                  '--']:
                self.dict_[key] = float(self.dict_[key])
        if '引擎' in self.dict_.keys() and isinstance(self.dict_['引擎'], str):
            if self.dict_['引擎'].lower() == 'aisinger':
                self.dict_['引擎'] = 'AISingers'
            elif self.dict_['引擎'].lower() == 'muta':
                self.dict_['引擎'] = 'MUTA'
            elif self.dict_['引擎'].lower() == 'dv':
                self.dict_['引擎'] = 'DeepVocal'
            elif self.dict_['引擎'].lower() == 'sk':
                self.dict_['引擎'] = 'Sharpkey'
            elif self.dict_['引擎'] == '袅袅':
                self.dict_['引擎'] = '袅袅虚拟歌手'
            elif self.dict_['引擎'].lower() == 'xstudio' or self.dict_['引擎'].lower() == 'x studio':
                self.dict_['引擎'] = 'X Studio'
            elif self.dict_['引擎'] == '其他':
                self.dict_['引擎'] = '其他/跨引擎'
        # self.dict_['pub_time'] = self.pub_time_datetime
        if str(self.dict_.get('原创')) == '1':
            self.dict_['原创'] = '原创'
        if self['原创'] == '榜外原创':
            self.dict_['原创'] = '原创'  # 榜外原创不需要额外标注，当期榜外原创由当期收录程序计算
        if str(self['原创']) == '0':
            self.dict_['原创'] = '其他'
        if self['主榜'] == '主榜截止':
            Data.main_flag = 1
        elif self['主榜'] == '副榜截止':
            Data.main_flag = 0

    def write_to_csv(self,
                     file_name: str,
                     head: list[str],
                     with_seconds: bool = False) -> None:
        #  续写，分条写入
        @permission_access_decorator
        def _write(file_name: str):
            nonlocal self
            dict_to_write = {key: value for key, value in self.dict_.items() if key in head}
            if with_seconds:
                dict_to_write['投稿时间'] = self.pub_time_datetime.strftime('%Y/%m/%d %H:%M:%S')
            else:
                dict_to_write['投稿时间'] = self.pub_time_datetime.strftime('%Y/%m/%d %H:%M')
            with open(file_name, 'a+', encoding='utf-8-sig', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=head)
                writer.writerow(dict_to_write)

        _write(file_name)

    def is_new(self) -> bool:
        if self['新曲'] in ['新曲榜', '新曲']:
            return True
        if '上次' in self.dict_.keys():
            return self['上次'] == 'NEW'
        return False

    def xlsx_cell_style(self) -> Tuple[op.styles.Font, op.styles.PatternFill, op.styles.Border]:
        font = op.styles.Font(name='等线', size=11)
        pattern = op.styles.PatternFill()
        border = op.styles.Border()
        if self['长期入榜及期数'] != '':
            pattern = op.styles.PatternFill(fill_type='solid', fgColor='FFC000', start_color='FFC000',
                                            end_color='FFC000')
            font.bold = True
        if self['主榜'] in ['主榜截止', '副榜截止']:
            border.bottom = op.styles.Side(style='medium', color='000000')
        if self['HOT'] == 'HOT':
            pattern = op.styles.PatternFill(fill_type='solid', fgColor='7030A0', start_color='7030A0',
                                            end_color='7030A0')
            font.color = op.styles.Color(rgb='FFFFFF')
            font.bold = True
        if not self.is_new():
            return font, pattern, border
        if self['原创'] == '原创' and self.is_new():
            pattern = op.styles.PatternFill(fill_type='solid', fgColor='C00000', start_color='C00000',
                                            end_color='C00000')
            font.color = op.styles.Color(rgb='FFFFFF')
            font.bold = True
        if self['原创'] == '榜外原创':
            pattern = op.styles.PatternFill(fill_type='solid', fgColor='92D050', start_color='92D050',
                                            end_color='92D050')
            font.bold = True
        if self['未授权搬运'] == '未授权搬运':
            pattern = op.styles.PatternFill(fill_type='solid', fgColor='FFC000', start_color='FFC000',
                                            end_color='FFC000')
            font.bold = True
        if self['收录'] == 0:
            pattern = op.styles.PatternFill(fill_type='solid', fgColor='FFFF00', start_color='FFFF00',
                                            end_color='FFFF00')
        return font, pattern, border

    def add_info(self, other, key=None):  # 添加信息，只会添加为没有或为空值的键值，不改变已有的键值
        if not self.is_same_song(other):
            print('aid不一致,不能合并')
            input()
            raise ValueError
        if key is None:
            for _key in other.dict_.keys():
                if _key not in self.dict_.keys():
                    self[_key] = other[_key]
                else:
                    if self[_key] == '':
                        self[_key] = other[_key]
        else:
            if key not in self.dict_.keys():
                self[key] = other[key]
            else:
                if self[key] == '' and other[key] != '':
                    self[key] = other[key]
        return

    def __getitem__(self, item: str | int) -> Data_value_type:
        if isinstance(item, str) and item.lower() == 'staff':
            return self.dict_['staff']
        if item in ['名次', '排名']:
            return self.dict_['名次']
        if isinstance(item, str) and item.lower() == 'lastpt':
            return self.dict_['Last Pt']
        elif isinstance(item, str) and item in self.dict_.keys():
            return self.dict_[item]
        elif isinstance(item, int) and item in Data.header_digit_to_word.keys():
            return self.dict_[Data.header_digit_to_word[item]]
        else:
            print("未知的key:", item)
            input()
            raise ValueError

    def __setitem__(self, key: str | int, value: Data_value_type):
        if isinstance(key, str) and key.lower() == 'staff':
            self.dict_['staff'] = value
        if key in ['名次', '排名']:
            self.dict_['名次'] = value
        if isinstance(key, str) and key.lower() == 'lastpt':
            self.dict_['Last Pt'] = value
        elif isinstance(key, int) and key in Data.header_digit_to_word.keys():
            self.dict_[Data.header_digit_to_word[key]] = value
        else:
            self.dict_[key] = value

    def __len__(self):
        return len(self.dict_.keys())

    def __lt__(self, other):
        if (self['HOT'] != 'HOT' and other['HOT'] != 'HOT') or (self['HOT'] == 'HOT' and other['HOT'] == 'HOT'):
            if float(self['Pt']) != float(other['Pt']):
                return float(self['Pt']) < float(other['Pt'])
            elif '投稿时间' in self.dict_.keys() and '投稿时间' in other.dict_.keys():
                return self.pub_time_datetime < other.pub_time_datetime
            else:
                return True
        else:
            if self['HOT'] == 'HOT':
                return False
            else:
                return True

    def is_same_song(self, other):
        return int(self['aid']) == int(other['aid'])

    def __contains__(self, item):
        return item in self.dict_.keys()


D = TypeVar('D', bound=Data)
Data_type = Union[Data, D]


@permission_access_decorator
def read(file_path: str,
         *,
         class_type: Type[Data_type] = Data,
         max_rank: int = -1,
         inclusion_status: int = 0, # 若等于一则所有曲子默认为确认收录状态
         required_keys: list[str] = None) -> list[Data_type]:
    data_list: list[Data_type] = []
    Data.main_flag = 0
    if not issubclass(class_type, Data):
        print('class_type必须为Data的子类')
        input()
        raise ValueError
    if file_path.split('.')[-1] == 'csv':
        try:
            with open(file_path, 'r') as f:
                _read_csv(f, data_list, class_type, max_rank, required_keys, inclusion_status)

        except UnicodeDecodeError:
            with open(file_path, 'r', encoding="utf-8-sig") as f:
                _read_csv(f, class_type, data_list,  max_rank, required_keys, inclusion_status)
    elif file_path.split('.')[-1] == 'xlsx':
        wb = op.load_workbook(file_path, read_only=True)
        sheet = wb.active
        xlsx_order: list[str] = ['']
        for g in sheet[1]:
            xlsx_order.append(g.value)
        last_process = 0
        row_count = sheet.max_row
        for g, row in enumerate(sheet.rows):
            if g == 0:
                continue
            progress = int((g - 1) * 100 / row_count)
            if progress % 10 == 0 and progress != last_process:
                print(str(progress) + '%')
                last_process = progress
            try:
                new_data = class_type(row, 'xlsx', xlsx_order, required_keys=required_keys)
                if new_data.valid:
                    if inclusion_status == 1:
                        new_data['收录'] = 1
                    data_list.append(new_data)
                    if (str(new_data['名次']) == str(max_rank)) and (max_rank != -1):
                        break
                else:
                    print('第' + str(g + 1) + '行数据非法，跳过')
            except ValueError as e:
                print(e)
                print('第' + str(g + 1) + '行出错')
                continue
    else:
        print("文件名格式错误")
        input()
        raise RuntimeError
    print("读取完成")
    return data_list


def _read_csv(f, class_type, data_list, max_rank, required_keys, inclusion_status):
    reader = csv.DictReader(f)
    for g, i in enumerate(reader):
        progress = int(g * 100 / reader.line_num)
        if progress % 10 == 0:
            print(f'{progress}%')
        try:
            new_data = class_type(i, 'csv', required_keys=required_keys)
            if new_data.valid:  # 避免空行
                if inclusion_status == 1:
                    new_data['收录'] = 1 if new_data['收录'] != 0 else 0
                data_list.append(new_data)
                if (str(new_data['名次']) == str(max_rank)) and (max_rank != -1):
                    break
            else:
                print('第' + str(g + 1) + '行数据非法，跳过')
        except ValueError as e:
            print(f'第{g + 1}行读取错误')
            print(e)
            continue


rank_trans = {0: "C", 1: "SV", 2: "U"}


def _input(text: str, valid: Callable[[str], bool], default=None):
    result = input(text) or default
    while result is None or not valid(result):
        print('输入格式错误')
        result = input(text)
    return result
