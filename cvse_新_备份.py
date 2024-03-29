# coding=utf-8
import csv
import webbrowser

import dateutil.relativedelta
import openpyxl as op

import collect_staff
import match
from downloader import *
from 主榜_包装 import generate
from 副榜_包装 import side_generate, move_file

DEBUG = False
header = ['名次', '上次', 'aid', '标题', 'mid', 'up主', '投稿时间', '时长', '分P数', '播放增量', '弹幕增量', '评论增量', '收藏增量', '硬币增量', '分享增量',
          '点赞增量', 'Pt', '修正A', '修正B', '修正C', '长期入榜及期数', '收录', "引擎", '原创', "主榜", 'Last Pt',
          'rate', 'staff', '新曲排名', '新曲', '未授权搬运', '已删稿', 'HOT']
engine = {1: 'Sharpkey', 2: 'DeepVocal', 3: 'MUTA', 4: '袅袅虚拟歌手', 5: 'AISingers', 6: 'X Studio', 7: '跨引擎'}
flag = 0
rank_trans = {0: "C", 1: "SV", 2: "U"}
# max_main = {0: 20, 1: 25}
# max_side = {0: 80, 1: 105}
# new_rank_number: dict[int, int] = {0: 10, 1: 8}
try:
    with open("config_inclusion.ini", 'r', encoding='utf-8') as f:
        config = json.load(f)
        with_match: int = int(config['with_match'])
        with_staff: int = int(config['with_staff'])
        with_template_generate: int = int(config['with_template_generate'])
        max_main: dict[int, int] = {int(i): j for i, j in config['max_main'].items()}
        max_side: dict[int, int] = {int(i): j for i, j in config['max_side'].items()}
        new_rank_number: dict[int, int] = {int(i): j for i, j in config['new_rank_number'].items()}
        min_duration: dict[int, int] = {int(i): j for i, j in config['min_duration'].items()}
except Exception as e:
    print(e)
    print("配置错误")
    input()
    exit()
if os.path.exists("remove.txt"):
    with open("remove.txt", 'r+') as f:
        remove_list = f.read().split('\n')
else:
    remove_list = []


def _input(text: str, valid, default=None):
    result = input(text) or default
    while result is None or not valid(result):
        print('输入格式错误')
        result = input(text)
    return result


def calculate_time(_rank: int, _index: int) -> (datetime.datetime, datetime.datetime):
    if _rank == 0:  # 国产榜
        basic_time_start = datetime.datetime.strptime("2021/04/28 3:00", "%Y/%m/%d %H:%M")
        basic_time_end = datetime.datetime.strptime("2021/05/28 3:00", "%Y/%m/%d %H:%M")
        time_start = basic_time_start + dateutil.relativedelta.relativedelta(months=_index - 48)
        time_end = basic_time_end + dateutil.relativedelta.relativedelta(months=_index - 48)
        return time_start, time_end
    if _rank == 1:  # SV刊
        basic_time_start = datetime.datetime.strptime("2021/11/26 3:00", "%Y/%m/%d %H:%M")
        basic_time_end = datetime.datetime.strptime("2021/12/03 3:00", "%Y/%m/%d %H:%M")
        time_start = basic_time_start + datetime.timedelta(weeks=_index - 132)
        time_end = basic_time_end + datetime.timedelta(weeks=_index - 132)
        return time_start, time_end


def calculate_index(_rank: int, time_start: datetime.datetime) -> int:
    if _rank == 0:
        temp = dateutil.relativedelta.relativedelta(dt1=time_start, dt2=datetime.datetime.strptime("2021/04/28 3:00","%Y/%m/%d %H:%M"))
        return temp.years * 12 + temp.months + 48
    if _rank == 1:
        temp = time_start - datetime.datetime.strptime("2021/11/26 3:00","%Y/%m/%d %H:%M")
        return temp.days // 7 + 132

def tag_info_decorator(func):
    # 实现对象临时存储简介和tag信息，若为空则下载相应信息，函数执行完成后复原，被装饰的函数可以直接调用正确（非空）的desc和tag，同时保证调用前后不会改变存储状态
    def wrapper(self, *args, **kwargs):
        init_tag = self.tag
        if not self.tag:
            res: dict = json.loads(
                request('https://api.bilibili.com/x/web-interface/view/detail/tag?aid=' + str(self['aid'])).text)
            self.tag = [i["tag_name"] for i in res['data']]
        result = func(self, *args, **kwargs)
        self.tag = init_tag
        return result

    return wrapper


def desc_title_info_decorator(func):
    # 实现对象临时存储简介和tag信息，若为空则下载相应信息，函数执行完成后复原，被装饰的函数可以直接调用正确（非空）的desc和tag，同时保证调用前后不会改变存储状态
    def wrapper(self, *args, **kwargs):
        init_desc = self.desc
        init_title = self.title
        if not self.desc:
            res_data = \
                json.loads(request('https://api.bilibili.com/x/web-interface/view?aid=' + str(self['aid'])).text)[
                    'data']
            if 'data' in res_data:
                self.desc = res_data['desc']
                self.title = res_data['title']
        result = func(self, *args, **kwargs)
        self.title = init_title
        self.desc = init_desc
        return result

    return wrapper


class Pres_data(CVSE_Data.Data):  # 添加新曲判断及收录判断
    xlsx_order: list[str] = [0]  # xlsx第i列列索引，从1开始
    flag: int = 0  # 是否完成收录
    index: int = 0  # 期数
    rank: int = -1
    start_time: datetime.datetime = None
    end_time: datetime.datetime = None
    max_count_main: int = 0  # 主榜最大曲数
    max_count_side: int = 0  # 副榜最大曲数
    min_duration: int = 0   # 最短时长，单位为秒
    remove_flag = 0

    @staticmethod
    def write_long_term_data_wrapper():
        wb = op.load_workbook(f'data_{rank_trans[Pres_data.rank]}.xlsx')
        ws = wb.active
        col = 2
        while ws.cell(1, col - 1).value != f'#{Pres_data.index - 1}':
            col += 1
            if col > 500:
                print(f'data_{rank_trans[Pres_data.rank]}.xlsx 格式错误')
        ws.cell(1, col).value = f'#{Pres_data.index}'
        Hot_line = 2
        line = 2
        while ws.cell(line, 1).value == 'HOT':
            line += 1

        def write_long_term_data(self: Pres_data) -> bool:  # 写入长期数据,返回主榜是否写完
            nonlocal ws
            nonlocal col
            nonlocal line
            nonlocal Hot_line
            if self['收录'] != 1:
                return True
            if self['HOT'] == 'HOT':
                ws.cell(Hot_line, col).value = int(self['aid'])
                Hot_line += 1
            elif str(ws.cell(line, 1).value) != str(self['名次']):
                print(f'{self["名次"]} 不在 {line} 行')
                raise Exception
            else:
                ws.cell(line, col).value = int(self['aid'])
                line += 1
            if self['主榜'] == '主榜截止':
                return False
            return True

        @CVSE_Data.permission_access_decorator
        def save_close(file_name: str):
            nonlocal wb
            nonlocal ws
            wb.save(file_name)

        return write_long_term_data, save_close

    def __init__(self, data, data_type: str, file_header: list = None):
        super(Pres_data, self).__init__(data, data_type, file_header)
        self.desc: str = ""
        self.tag: list[str] = []
        self.title: str = ''
        self['上次'] = '——'
        if flag and self.dict_['收录'] == '':
            self.dict_['收录'] = 1
        if (not flag) and self.dict_['收录'] == '':
            self.dict_['收录'] = ''  # 待收录
        if '投稿时间' in self.dict_.keys() and self.dict_['投稿时间'] != '':
            if Pres_data.end_time > self.pub_time_datetime > Pres_data.start_time:
                self.dict_['新曲'] = '新曲'
                self['上次'] = 'NEW'
            elif self.pub_time_datetime > Pres_data.end_time:
                self.dict_['新曲'] = ''
                self.dict_['收录'] = 0
            else:
                self.dict_['新曲'] = ''
                self.dict_['收录'] = 1
        if str(self.dict_['aid']) in remove_list:
            self.dict_['收录'] = 0
        if 0 in self.dict_.keys():
            del self.dict_[0]

    def duration_check(self) -> int:
        # 检查时长, -1表示已删稿
        res2 = request('https://api.bilibili.com/x/player/pagelist?aid=' + self['aid'])
        res2 = json.loads(res2.text)
        if not "data" in res2:
            return -1
        flag = 0
        for i in res2["data"]:
            if int(i.get("duration")) >= Pres_data.min_duration:
                flag = 1
                break
        return flag
    @desc_title_info_decorator
    def get_staff(self, with_open_browser: bool = False):
        if not with_staff:
            return
        if self['staff'] != '':
            return
        if with_open_browser:
            av = self['aid']
            webbrowser.open("https://www.bilibili.com/video/av" + str(av))
        _staff = ''
        if self['引擎'] == '':
            if Pres_data.rank == 0:
                self['引擎'] = _input("引擎为：1=SK 2=DV 3=Muta 4=袅袅 5=AiSinger 6=Xstudio 7=跨引擎\n",
                                    lambda x: x.isdigit() and int(x) in list(range(1, 8)))
                self['引擎'] = engine[int(self['引擎'])]
                _staff += f"{self['引擎']}  |  "
            else:
                self['引擎'] = rank_trans[Pres_data.rank]
        elif Pres_data.rank == 0:
            _staff += f"{self['引擎']}  |  "
        if self['原创'] == '':
            ori = _input("1=原创 2=未授权搬运（授权搬运不用标注） 3=其他\n", lambda x: str(x) in ['1', '2', '3'])
            if ori == '1':
                self['原创'] = '原创'
            else:
                self['原创'] = '其他'
            if ori == '2':
                self['未授权搬运'] = '未授权搬运'
        if self['原创'] == '其他' or self['原创'] == '':
            ori_work = input('原作:')
            _staff += '原作:' + ori_work + '  |  '
        # res2 = request('https://api.bilibili.com/x/web-interface/view?aid=' + str(self['aid']))
        # res2 = json.loads(res2.text)
        # staff = collect_staff.Staff(res2['data']['desc'], 1).staff_dict_degeneracy_str
        staff = collect_staff.Staff(self.desc, 1).staff_dict_degeneracy_str
        confirm = 'n'
        new_staff = ''
        while confirm == 'n':
            new_staff = input('输入staff,自动生成为：' + staff + '  可回车确定或输入新staff\n') or staff
            confirm = _input("确认staff y/n\n", lambda x: x in ['y', 'n'])
        _staff += new_staff
        self['staff'] = _staff

    def inclusion(self, place: int, new_place: int):  # 收录, 参数分别为此曲排名和此曲作为新曲的排名, 返回值为下一位的排名和作为新曲的排名
        def staff_info_confirm(browser_flag):
            nonlocal self, place, new_place
            if place <= Pres_data.max_count_main and self['staff'] == '':
                if with_staff:
                    self.get_staff(not browser_flag)
                    if self['原创'] == '其他':
                        self['原创'] = ''
            if self.is_new() and place > Pres_data.max_count_main and new_place <= new_rank_number[
                Pres_data.rank]:
                self['新曲'] = '新曲榜'
                if with_staff:
                    self.get_staff(not browser_flag)
                    if self['原创'] == '其他':
                        self['原创'] = ''
            if place > Pres_data.max_count_main and self['原创'] == '原创':
                self['原创'] = '榜外原创'

        def rank_info_confirm():
            nonlocal self, place, new_place
            if place == Pres_data.max_count_main and self['HOT'] != 'HOT' and self['长期入榜及期数'] == '':
                self['主榜'] = '主榜截止'
            elif place == Pres_data.max_count_side:
                self['主榜'] = '副榜截止'
            else:
                self['主榜'] = ''
            self['名次'] = place if self['HOT'] != 'HOT' else 'HOT'
            self['新曲排名'] = new_place if self['新曲'] != '' else 0

        #@tag_info_decorator
        #@desc_title_info_decorator
        # 本来设想通过tag和简介提取一些引擎关键词，但是因为网速问题可能体验不太好略显鸡肋，待定
        def info_input(self: Pres_data):
            # 对于新曲需要标题,简介和tag信息用于更新引擎原创及staff，这里把所有需要的部分包装起来了
            nonlocal place, new_place
            # print(self.title)
            # print(self.desc)
            # print(self.tag)
            if Pres_data.rank == 0:
                #  国产榜
                self['引擎'] = _input("引擎为：1=SK 2=DV 3=Muta 4=袅袅 5=AiSinger 6=Xstudio 7=跨引擎\n",
                                    lambda x: x.isdigit() and int(x) in list(range(1, 8)))
                self['引擎'] = engine[int(self['引擎'])]
            else:
                self['引擎'] = rank_trans[Pres_data.rank]
            ori = _input("1=原创 2=未授权搬运（授权搬运不用标注） 3=其他\n", lambda x: str(x) in ['1', '2', '3'])
            if ori == '1':
                self['原创'] = '原创'
            else:
                self['原创'] = '其他'
            if ori == '2':
                self['未授权搬运'] = '未授权搬运'
            rank_info_confirm()
            staff_info_confirm(browser_flag=True)

        res2 = 0
        if self.is_new():
            self['上次'] = 'NEW'
        hot_flag = False
        browser_flag = 0  # 是否已打开浏览器
        av = self['aid']
        if DEBUG:
            self['收录'] = 1
            self['staff'] = '23132312'
        if self['收录'] == '':
            res: int = self.duration_check()
            if res == -1:
                self['已删稿'] = 1
                self['收录'] = 0
                Pres_data.remove_flag = 1
                return place, new_place
            if res == 0:
                self['收录'] = 0
                Pres_data.remove_flag = 1
                return place, new_place
            webbrowser.open("https://www.bilibili.com/video/av" + str(av))
            browser_flag = 1
            inclusion = _input("是否收录 y/n，默认为y\n", lambda x: x in ['y', 'n'], 'y')
            if inclusion == 'n':
                self['收录'] = 0
                remove_list.append(self['aid'])
                Pres_data.remove_flag = 1
                return place, new_place
            self['收录'] = 1
            info_input(self)  # 这里已经包含了下面的rank_info_confirm和staff_info_confirm两个方法，这么包装只是为了保证不会把简介和tag信息请求两次
        elif self['收录'] == 1:
            if self['HOT'] == '两次前三' and place <= 3:
                self['HOT'] = 'HOT'
                self['名次'] = 'HOT'
                hot_flag = True
            elif self['HOT'] == 'HOT':
                self['名次'] = 'HOT'
                hot_flag = True
            elif place <= Pres_data.max_count_main and self['长期入榜及期数'] != '':
                Pres_data.max_count_main += 1
                Pres_data.max_count_side += 1
            rank_info_confirm()
            staff_info_confirm(browser_flag=browser_flag)
        elif self['收录'] == 0:
            Pres_data.remove_flag = 1
            return place, new_place
        else:
            raise ValueError
        if with_template_generate:
            if place <= Pres_data.max_count_main:
                download_cover(str(av), 'cover/AV' + str(av) + '.jpg')
                download_face(str(self['mid']), 'side_cover/uid' + str(self['mid']) + '-' + self['up主'] + '.jpg')
            if place <= Pres_data.max_count_side or self.is_new():
                download_cover(str(av), 'side_cover/AV' + str(av) + '.jpg')
                download_face(str(self['mid']), 'side_cover/uid' + str(self['mid']) + '-' + self['up主'] + '.jpg')
            if new_place <= new_rank_number[Pres_data.rank] and self.is_new():
                download_face(str(self['mid']), 'side_cover/uid' + str(self['mid']) + '-' + self['up主'] + '.jpg')
        if self['新曲'] != '':
            return place + 1, new_place + 1
        elif hot_flag:
            return place, new_place
        else:
            return place + 1, new_place


class Rank_data:
    data_list = ['播放', '弹幕', '评论', '收藏', '硬币', '分享', '点赞']
    engine_list = ['袅袅虚拟歌手', 'MUTA', 'Sharpkey', 'DeepVocal', 'AISingers', 'X Studio', '跨引擎']

    def __init__(self, data: list[CVSE_Data.Data]):
        self.Data_list = data
        self.count = {i: 0 for i in Rank_data.data_list + ['新曲']}  # 只需计数
        for i in Rank_data.engine_list + ['原创']:
            self.count[i] = 0
        self.aid_list = {i: [] for i in Rank_data.engine_list + ['原创']}  # 需减去上期
        for i in self.Data_list:
            if i['收录'] == 0:
                continue
            for j in Rank_data.data_list:
                self.count[j] += i[j + '增量']
            if i['原创'] in ['原创', '榜外原创']:
                self.aid_list['原创'] += [i['aid']]
                self.count['原创'] += 1
            if i.is_new():
                self.count['新曲'] += 1
            if i['引擎'] in Rank_data.engine_list:
                self.count[i['引擎']] += 1
                self.aid_list[i['引擎']].append(i['aid'])


class Rank_data_delta:
    def __init__(self, pres_data: Rank_data, prev_data: Rank_data):
        self.data_delta = {}
        self.data_new = {}
        for key in Rank_data.data_list + ['新曲']:
            self.data_delta[key] = pres_data.count[key] - prev_data.count[key]
        for key in Rank_data.engine_list + ['原创']:
            pres_new_count = [1 for i in pres_data.aid_list[key] if i not in prev_data.aid_list[key]]
            self.data_delta[key] = sum(pres_new_count) - prev_data.count[key]
            self.data_new[key] = sum(pres_new_count)
        self.data_delta['其他/跨引擎'] = self.data_delta['跨引擎']
        self.data_new['其他/跨引擎'] = self.data_new['跨引擎']


def to_str_with_delta(pres_data: Rank_data, prev_data: Rank_data):
    data_delta = Rank_data_delta(pres_data, prev_data)
    key_list_1 = Rank_data.data_list + ['新曲']
    key_list_2 = Rank_data.engine_list + ['原创']
    key_list_2[-2] = '其他/跨引擎'
    sign = lambda key: '+' if data_delta.data_delta[key] >= 0 else ''
    temp_list = [f'{i}:{pres_data.count[i]}({sign(i)}{data_delta.data_delta[i]})' for i in key_list_1]
    temp_list += [f'{i}:{data_delta.data_new[i]}({sign(i)}{data_delta.data_delta[i]})' for i in key_list_2]
    out_str = '\n'.join(temp_list)
    return out_str


def init() -> tuple[list[Pres_data], int, int, str]:
    global flag
    rank = input("请输入国产榜/SV/utau（还没写），1=国产榜，2=SV刊\n")
    while rank not in ['1', '2']:
        print('输入错误')
        rank = input("请输入国产榜/SV/utau（还没写），1=国产榜，2=SV刊\n")
    rank = int(rank) - 1
    Pres_data.rank = rank
    Pres_data.max_count_main = max_main[rank]
    Pres_data.max_count_side = max_side[rank]
    Pres_data.min_duration = min_duration[rank]
    index = _input("请输入待处理排行榜期数，如 133 50\n", lambda x: x.isdigit())
    while not index.isdigit():
        print('输入的不是数字')
        index = input("请输入待处理排行榜期数，如 133 50\n")
    index = int(index)
    Pres_data.index = index
    Pres_data.start_time, Pres_data.end_time = calculate_time(int(rank), int(index))
    default_dir = f'{rank_trans[rank]}_{index}'
    text: str = f'请输入待处理文件名，如 synthv增量_220304.csv  133.xlsx, 文件格式只限csv和xlsx 将在当前目录和当前目录下的{default_dir}搜索\n输入1自动下载原始数据文件或读取已下载的原始数据文件'
    if os.path.exists(f'{default_dir}/{rank_trans[rank]}_{index}_save_backup.csv'):

        file = input(f'请输入待处理文件名，如 synthv增量_220304.csv  133.xlsx, 文件格式只限csv和xlsx 将在当前目录和当前目录下的{default_dir}搜索, '
                     f'默认为{rank_trans[rank]}_{index}_save_backup.csv\n') or f'{rank_trans[rank]}_{index}_save_backup.csv'
    else:
        file = input(text)
    while not os.path.exists(file):
        if file == '1':
            if not os.path.exists(default_dir):
                os.mkdir(default_dir)
            status: int = download_pres_data(Pres_data.end_time, rank, index, default_dir)
            if status:
                file = f'{default_dir}/{index}.csv'
                break
        if os.path.exists(f'{default_dir}/' + file):
            file = f'{default_dir}/' + file
            break
        print('文件不存在')
        file = input(text)
    flag = _input('待处理文件是否已经完成收录？y/n\n', lambda x: x in ['y', 'n'])
    if flag == 'y':
        flag = 1
    else:
        flag = 0
    data_list: list[Pres_data] = CVSE_Data.read(file, Pres_data)
    return data_list, rank, index, default_dir


def read_last(rank: int, index: int):
    # index是上一期的序号
    default_dir = f'{rank_trans[rank]}_{index}'
    if os.path.exists(f'{default_dir}/{rank_trans[rank]}_{index}_save_backup.csv'):
        file = input(
            f'请输入上期排行榜的文件名，如 synthv增量_220304.csv  133.xlsx, 文件格式只限csv和xlsx 将在当前目录和当前目录下的{default_dir}搜索'
            f'默认为{rank_trans[rank]}_{index}_save_backup.csv\n') or f'{rank_trans[rank]}_{index}_save_backup.csv'
    else:
        text: str = f'请输入上期排行榜的文件名，如 synthv增量_220304.csv  133.xlsx, 文件格式只限csv和xlsx 将在当前目录和当前目录下的{default_dir}搜索\n' + f'输入1自动下载往期数据文件或读取已下载的数据文件 '
        file: str = input(text)
    while not os.path.exists(file):
        if file == '1':
            if not os.path.exists(default_dir):
                os.mkdir(default_dir)
            _, prev_end_time = calculate_time(rank, index)
            status: int = download_history_data(prev_end_time, rank, index, default_dir)
            if status:
                file = f'{default_dir}/{index}.xlsx'
                break
        if os.path.exists(f'{default_dir}/' + file):
            file = f'{default_dir}/' + file
            break
        print('文件不存在')
        file = input(f'请输入上期排行榜的文件名，如 synthv增量_220304.csv  133.xlsx, 文件格式只限csv和xlsx 将在当前目录和当前目录下的{default_dir}搜索\n')
    return CVSE_Data.read(file, CVSE_Data.Data)


def load_record(_pres_list: list[CVSE_Data.Data], file_name: str):
    if not os.path.exists(file_name):
        return _pres_list
    print("读取记录")
    record_list = CVSE_Data.read(file_name, Pres_data)
    temp_dict = {int(i['aid']): i for i in record_list}
    [i.add_info(temp_dict[int(i['aid'])]) for i in _pres_list if int(i['aid']) in temp_dict.keys()]
    return


def history(rank: int, index: int):
    # index是当前期的序号
    default_dir: str = f'{rank_trans[rank]}_{index}'
    his_index: int = calculate_index(rank, Pres_data.start_time - dateutil.relativedelta.relativedelta(years=1))
    text: str = f'请输入历史回顾当期（{his_index}期）数据的文件名，如 41-2010.xlsx, 文件格式只限csv和xlsx 将在当前目录和当前目录下的{default_dir}搜索\n输入1自动下载历史数据文件或读取已下载的数据文件，输入0跳过\n'
    file: str = input(text)
    while not os.path.exists(file):
        if file == '0':
            return
        if file == '1':
            status: int = download_history_data(Pres_data.end_time - dateutil.relativedelta.relativedelta(years=1), Pres_data.rank, his_index, default_dir)
            if status:
                file = f'{default_dir}/{his_index}.xlsx'
                break
        if os.path.exists(f'{default_dir}/' + file):
            file = f'{default_dir}/' + file
            break
        print('文件不存在')
        file = input(text)
    his_data = CVSE_Data.read(file, CVSE_Data.Data, 5)
    write_xlsx, save = CVSE_Data.Data.write_to_xlsx_wrapper(header=CVSE_Data.history_header)
    for i in his_data:
        write_xlsx(i)
    his_cover_file = f"cover/history.jpg"
    if not os.path.exists(his_cover_file):
        his_bv = input(f"请输入历史回顾排行榜(第{his_index})期的aid/bvid")
        download_cover(his_bv, his_cover_file)
    save(f'{_default_dir}/{rank_trans[_rank]}_{_index}_历史.xlsx')
    if with_template_generate:
        generate(his_data, 'history', CVSE_Data.history_header, out_path=os.path.join(default_dir, '模板'))
    return


def pick_up(pres_data: list[Pres_data], rank: int, index: int):
    default_dir: str = f'{rank_trans[rank]}_{index}'
    pk_aid = int(_input("请输入pick up视频的aid，一次输入一个，输入0退出", lambda x: x.isdigit()))
    if pk_aid != 0:
        write_xlsx_pk, save_pk = CVSE_Data.Data.write_to_xlsx_wrapper()
        pk_data: list[Pres_data] = []
        while pk_aid != 0:
            temp_list: list[Pres_data] = [i for i in pres_data if int(i['aid']) == pk_aid]
            if not temp_list:
                print('aid错误')
                continue
            data: Pres_data = temp_list[0]
            data.get_staff(with_open_browser=True)
            pk_aid = int(_input("请输入pick up视频的aid，一次输入一个，输入0退出", lambda x: x.isdigit()))
        pk_data.sort(reverse=True)
        for i in pk_data:
            write_xlsx_pk(i)
        save_pk(f'{_default_dir}/{rank_trans[_rank]}_{_index}_pick_up.xlsx')
        if with_template_generate:
            generate(pk_data, 'pick_up', CVSE_Data.xlsx_header, out_path=os.path.join(default_dir, '模板'))


for i in ['cover', 'side_cover']:
    if not os.path.exists(i):
        os.mkdir(i)
pres_list, _rank, _index, _default_dir = init()
if with_match:
    prev_list = read_last(_rank, int(_index) - 1)
    print('开始与上期数据进行匹配')
    match.match(_rank, _index, Pres_data.start_time, pres_list, prev_list)
    print('匹配完成')
if Pres_data.flag != 1:
    load_record(pres_list, f'{_default_dir}/{rank_trans[_rank]}_{_index}_save.csv')
    load_record(pres_list, f'{_default_dir}/{rank_trans[_rank]}_{_index}_save_backup.csv')
with open(f'{_default_dir}/{rank_trans[_rank]}_{_index}_save.csv', 'w', newline='', encoding='utf-8-sig') as f:
    f = csv.DictWriter(f, fieldnames=header)
    f.writeheader()
place = 1
new_place = 1
pres_list.sort(reverse=True)
new_rank_list: list[Pres_data] = []
with open(f'{_default_dir}/remove_{_index}.txt', 'w+') as remove_pres:
    for i in pres_list:
        place, new_place = i.inclusion(place, new_place)
        if i['新曲'] == '新曲榜':
            new_rank_list.append(i)
        i.write_to_csv(f'{_default_dir}/{rank_trans[_rank]}_{_index}_save.csv', header)
        if Pres_data.remove_flag:
            with open('remove.txt', 'a+') as remove_f:
                remove_f.write(str(i['aid']) + '\n')
            Pres_data.remove_flag = 0
        if i.is_new() and i['收录'] == 0:
            remove_pres.write(str(i['aid']) + '\n')
pres_list.sort(reverse=True)
write_xlsx, save = CVSE_Data.Data.write_to_xlsx_wrapper()
write_xlsx_, save_ = CVSE_Data.Data.write_to_xlsx_wrapper()
outfile_header = ['名次', '上次', 'aid', '标题', 'mid', 'up主', '投稿时间', '时长', '分P数', '播放增量', '弹幕增量', '评论增量', '收藏增量', '硬币增量',
                  '分享增量',
                  '点赞增量', 'Pt', '修正A', '修正B', '修正C', 'Last Pt', 'rate', '长期入榜及期数', '新曲排名']
with open(f'{_default_dir}/{rank_trans[_rank]}_{_index}_save_backup.csv', 'w', newline='', encoding='utf-8-sig') as f:
    f = csv.DictWriter(f, fieldnames=header)
    f.writeheader()
if with_template_generate:
    with open(f'outfile.csv', 'w', newline='', encoding='utf-8-sig') as f:
        f = csv.DictWriter(f, fieldnames=outfile_header)
        f.writeheader()
for i in pres_list:
    i.write_to_csv(f'{_default_dir}/{rank_trans[_rank]}_{_index}_save_backup.csv', header)
    if with_template_generate:
        if str(i['收录']) != '0' and i['HOT'] != 'HOT':
            i.write_to_csv(f'outfile.csv', outfile_header)
for idx, i in enumerate(pres_list):
    write_xlsx(i)
    if i['收录'] != 0:
        write_xlsx_(i)
    if idx % 100 == 0:
        print(f'正在写入第{idx}条数据')
save(f'{_default_dir}/{rank_trans[_rank]}_{_index}_含不收录曲.xlsx')
save_(f'{_default_dir}/{rank_trans[_rank]}_{_index}.xlsx')
print(f'已保存为{_default_dir}/{rank_trans[_rank]}_{_index}.xlsx')
if new_rank_list:
    write_new_xlsx, save_new = CVSE_Data.Data.write_to_xlsx_wrapper()
    for i in new_rank_list:
        write_new_xlsx(i)
    save_new(f'{_default_dir}/{rank_trans[_rank]}_{_index}_新曲榜.xlsx')
if with_match:
    write_long_term_xlsx, save = Pres_data.write_long_term_data_wrapper()
    i = iter(pres_list)
    while write_long_term_xlsx(i.__next__()):
        pass
    save(f'data_{rank_trans[Pres_data.rank]}.xlsx')
    pres_rank_data = Rank_data(pres_list)
    prev_rank_data = Rank_data(prev_list)
    rank_information = to_str_with_delta(pres_rank_data, prev_rank_data)
    print(rank_information)
    with open(f'{_default_dir}/{rank_trans[_rank]}_{_index}_数据信息.txt', 'w') as f:
        f.write(rank_information)
if with_template_generate:
    print('正在生成模板')
    history(_rank, _index)
    generate(pres_list, 'main', CVSE_Data.xlsx_header, out_path=os.path.join(_default_dir, '模板'), end_flag=('主榜', '主榜截止'))
    generate(new_rank_list, 'new_rank', CVSE_Data.xlsx_header, out_path=os.path.join(_default_dir, '模板'))
    trans = lambda x: int((3 * x ** 2 - 5 * x + 4) / 2)  # 只是转换一下两边的序号
    side_generate(trans(Pres_data.rank), Pres_data.max_count_main + 1, Pres_data.max_count_side)
    if not os.path.exists(f'{_default_dir}/模板'):
        os.mkdir(f'{_default_dir}/模板')
    move_file('side', f'{_default_dir}/模板')
    print('模板生成完成')

input('按任意键退出')
